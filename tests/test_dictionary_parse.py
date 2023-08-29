# Created by alanhu at 8/28/23
import unittest
import sys
import os
from pathlib import Path
import re
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
import pandas as pd

# Get the directory of the current script
current_script_directory = os.path.dirname(os.path.abspath(__file__))

# Get the parent directory
parent_directory = os.path.dirname(current_script_directory)

# Append the parent directory to sys.path
sys.path.append(parent_directory)


def save_excel_file(modified_data, headers_xlsx):
    # Create a new Excel workbook and worksheet for the processed data

        wb_processed = Workbook()
        ws_processed = wb_processed.active

        # Write headers to the worksheet
        ws_processed.append(headers_xlsx)

        # Write the processed data to the worksheet
        for row in modified_data:
            ws_processed.append(row)

        # Save the workbook to a new XLSX file

        xlsx_processed_file_path = parent_directory + "/ASIObjectDictionary_new.xlsx"
        wb_processed.save(xlsx_processed_file_path)


class MyTestCase(unittest.TestCase):
    def test_something(self):
        self.assertEqual(True, False)  # add assertion here

    def setUp(self):
        # Read the data from the Excel file
        # Load the newly provided XLSX file
        path_xlsx = parent_directory + "/ASIObjectDictionary.xlsx"

        wb_loaded = load_workbook(path_xlsx)
        ws_loaded = wb_loaded.active

        # Extract data from the worksheet
        data_from_xlsx = [row for row in ws_loaded.iter_rows(values_only=True)]

        self.data_from_xlsx = data_from_xlsx
        # Extract headers for further reference
    #    headers_xlsx = data_from_xlsx[0]
     #   data_from_xlsx = data_from_xlsx[1:]  # Removing header row for processing

     #   headers_xlsx, data_from_xlsx[:5]  # Displaying headers and first few rows for verification

    # Function to format bit information as per the given instructions
    def format_bit_info(self, base_row, bit_number, bit_name):
        # Format Address
        address = f"{base_row[0]}.{bit_number:02}"

        # Format Name
        name = bit_name.strip()

        # Extract other columns from the base row
        scale = base_row[2]
        units = base_row[3]
        access_level = base_row[6]

        # Format Description
        description = f"bit {bit_number}: {name}"

        # Format Key
        key = "_".join(re.sub(r"[^a-zA-Z0-9 ]", "", name).lower().split())

        # Bitfield
        bitfield = str(bit_number)

        return (address, name, scale, units, description, key, access_level, bitfield)

    # Function to modify the data as per the given instructions
    def test_modify_data(self):

        data_from_xlsx = self.data_from_xlsx

        modified_data = []
        for row in data_from_xlsx:
            modified_data.append(row)  # Append the original row
            if row[2] == "bit vector":
                # Extract bit information from the Description
                bits = re.findall(r"bit (\d+): ([^\n]+)", row[4])
                for bit_info in bits:
                    bit_number = int(bit_info[0])
                    bit_name = bit_info[1]
                    modified_data.append(self.format_bit_info(row, bit_number, bit_name))

        # Display the first few rows of the modified data for verification
        for row in modified_data:
            print(row)

        headers_xlsx = data_from_xlsx[0]

        save_excel_file(modified_data, headers_xlsx)

    # Display the first few rows of the modified data for verification
    # modified_data[:15]


if __name__ == '__main__':
    unittest.main()
